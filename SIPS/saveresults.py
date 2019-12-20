# -*- coding: utf-8 -*-
"""
Created on Thu Dec 19 16:55:40 2019

@author: ignac
"""
# from openpyxl import Workbook, load_workbook
# filename='d3.xlsx'
# wb = Workbook()
# sheet = wb.active
# sheet['A1'] = 12312
# wb.save(filename=filename)

from openpyxl import Workbook
import os
import xlwings as xw



# wb = xw.Book(filename)
# sht = wb.sheets['GEN']
# sht.range('A1').value = np.array([[1,2],[3,4]])

# wb.save()

def saveResults(psol,fsol,rupsol,rdownsol,xsol,startcol,batsol=None):
    filename = 'data.xlsx'
    curFol = os.path.realpath('')
    path_filename = curFol + '\\' + filename
    batcol = chr(ord(startcol)+1)
    rupcol = chr(ord(startcol)+2)
    rdowncol = chr(ord(startcol)+3)
    wb = xw.Book(path_filename)
    sht = wb.sheets['RES']
    sht.range(startcol+str(2)).value = psol
    sht.range(startcol+str(5)).value = fsol
    sht.range(rupcol+str(10)).value = rupsol[0]
    sht.range(rupcol+str(11)).value = rupsol[1]
    sht.range(rupcol+str(12)).value = rupsol[2]
    
    sht.range(rdowncol+str(10)).value = rdownsol[0]
    sht.range(rdowncol+str(11)).value = rdownsol[1]
    sht.range(rdowncol+str(12)).value = rdownsol[2]
    if batsol is not None:
        sht.range(batcol+str(10)).value = batsol[0]
        sht.range(batcol+str(11)).value = batsol[1]
        sht.range(batcol+str(12)).value = batsol[2]
    wb.save()
# book = Workbook()
# sheet = book.active

# rows = (
#     (88, 46, 10000),
#     (89, 38, 12),
#     (23, 59, 78),
#     (56, 21, 98),
#     (24, 18, 43),
#     (34, 15, 67)
# )

# for row in rows:
#     sheet.append(row)

# book.save(path_filename)